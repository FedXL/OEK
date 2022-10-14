from openpyxl import load_workbook
import pandas as pd
wb = load_workbook(filename="1s.xlsx")
sheet = wb.active



def sh():
    s = str(sheet)
    a=s[10:-1]
    return a



def settings ():
    #вычисляем параметры матрицы
    #return a,b
    #a= max_colums
    #b= max_rows
    i,j = 1,1
    a,b = 0,0
    while a != None:
        a = sheet.cell(row = 1, column = i).value
        i += 1
    while b != None:
        b = sheet.cell(row = j, column = 1).value
        j += 1
    return i-1,j-1



def frame ():
    a,b = settings()
    A = [[sheet.cell(row = j , column = i).value for i in range (2,a)] for j in range (1,b)]
    D = {}
    for i in range (0,len(A)):
        D[sheet.cell(row = i+1 , column = 1).value] = A[i]
    df = pd.DataFrame(D).T
    return df
