from openpyxl import load_workbook
import pandas
import progress
import time
from progress.bar import IncrementalBar
import numpy


wb = load_workbook(filename="КЛ.xlsx")
sheet = wb.active

def settings ():
    #вычисляем параметры матриц
    #return a,b
    #a= max_colums
    #b= max_rows
    i,j = 1,1
    a,b = 0,0
    while a != None:
        a = sheet.cell(row = 1, column = i).value
        i += 1
    while b != None:
        b = sheet.cell(row = j, column = 1).value
        j += 1
    return i-1,j-1

def frame ():
    a,b = settings()
    A = [[sheet.cell(row = j , column = i).value for i in range (2,a)] for j in range (1,b)]
    D = {}
    for i in range (0,len(A)):
        D[sheet.cell(row = i+1 , column = 1).value] = A[i]
    df = pandas.DataFrame(D).T
    return df


a = (frame())
def make_vocabulary(a):
    """Формирует словарь из перечня КЛ"""
    d = {}
    n = 0
    for i in a[1]:
        if d.get(i.replace(" ","")) == None:
            d[i.replace(" ","")] = []
            n+=1
        else:
            None
    return d
DIKT_3 = make_vocabulary(a)

def make_vocabulary_3(a,b):
    """ ВАЖНО все ключи словаря используем через убирание пробелов """
    per = ""
    for i in range (1,len(b)):
        per = b[1][i]
        if ("РЭС" in str(b[7][i])) or ("стадии" in str(b[7][i])):
            a[per.replace(" ","")].append((b[2][i],b[5][i],b[7][i]))
        else:
            None
make_vocabulary_3 (DIKT_3,a)
