from openpyxl import load_workbook
import pandas
import progress
import time
from progress.bar import IncrementalBar
import numpy

wb = load_workbook(filename="журнал.xlsx")
sheet = wb.active

def settings ():
    #вычисляем параметры матрицы
    #return a,b
    #a= max_colums
    #b= max_rows
    i,j = 1,1
    a,b = 0,0
    while a != None:
        a = sheet.cell(row = 1, column = i).value
        i += 1
    while b != None:
        b = sheet.cell(row = j, column = 1).value
        j += 1
    return i-1,j-1

def frame ():
    a,b = settings()
    A = [[sheet.cell(row = j , column = i).value for i in range (2,a)] for j in range (1,b)]
    D = {}
    for i in range (0,len(A)):
        D[sheet.cell(row = i+1 , column = 1).value] = A[i]
    df = pandas.DataFrame(D).T
    return df

def make_massiv (D):
    a = []
    j=0
    for i in D[1]:
        a.append(i.split("|"))
        a[j].pop(0)
        j+=1
    j=0
    for i in D[2]:
        a[j].append(i)
        j+=1
    j=0
    for i in D[3]:
        a[j].append(i)
        j+=1
    return a

a = make_massiv(frame())

def make_vocabulary(a):
    """ формирует словарь из журнала"""
    a.pop(0)
    d = {}
    for i in range (len(a)):
        if d.get(a[i][0]) == None:
            d[a[i][0]] = []
        else:
            None
    for i in range(len(a)):
        d[a[i][0]].append((a[i][1],str(a[i][2]).replace(" ",""),a[i][3]))
    return d



DIKT_2 = make_vocabulary(make_massiv(frame()))


pew = 0
# работа с файлом
for i in DIKT_2.keys():
    pew += 1
    print(pew,")",i)
